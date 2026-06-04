#!/usr/bin/env python3
import argparse
import json
import re
from collections import OrderedDict
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


def normalize_text(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    return str(value).strip()


def normalize_date(value):
    text = normalize_text(value)
    if not text:
        return ""

    if isinstance(value, (datetime, date)):
        return normalize_text(value)

    # Common date strings: 2026-05-12 / 2026/5/12 / 2026年5月12日 / 5月12日
    patterns = [
        r"(?P<y>\d{4})[-/.年](?P<m>\d{1,2})[-/.月](?P<d>\d{1,2})日?",
        r"(?P<m>\d{1,2})月(?P<d>\d{1,2})日",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if not match:
            continue
        parts = match.groupdict()
        year = int(parts.get("y") or datetime.now().year)
        month = int(parts["m"])
        day = int(parts["d"])
        try:
            return date(year, month, day).isoformat()
        except ValueError:
            return ""

    # Excel might store a date as a serial string in a text cell.
    try:
        parsed = datetime.fromisoformat(text.replace("/", "-"))
        return parsed.date().isoformat()
    except ValueError:
        return text


def unique_list(values):
    seen = OrderedDict()
    for value in values:
        text = normalize_text(value)
        if text and text not in seen:
            seen[text] = None
    return list(seen.keys())


def find_header_row(ws, search_limit=12):
    for row in range(1, min(search_limit, ws.max_row) + 1):
        for cell in ws[row]:
            if normalize_text(cell.value) == "序号":
                return row
    return None


def read_key_value_sheet(ws):
    mapping = {}
    for row in ws.iter_rows(min_row=1, values_only=True):
        key = normalize_text(row[0] if len(row) > 0 else None)
        value = row[1] if len(row) > 1 else None
        if not key or key in {"字段", "使用规则", "Sheet"}:
            continue
        mapping[key] = value
    return mapping


def read_table_sheet(ws):
    header_row = find_header_row(ws)
    if header_row is None:
        return []

    headers = []
    for cell in ws[header_row]:
        header = normalize_text(cell.value)
        if header:
            headers.append((cell.column, header))

    rows = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=False):
        values = {}
        has_data = False
        for col_idx, header in headers:
            cell_value = row[col_idx - 1].value if col_idx - 1 < len(row) else None
            text_value = normalize_text(cell_value)
            if text_value:
                has_data = True
            values[header] = cell_value
        if has_data:
            rows.append(values)
    return rows


def build_meta(workbook_path):
    wb = load_workbook(workbook_path, data_only=True)
    folder_name = Path(workbook_path).resolve().parent.name
    file_name = Path(workbook_path).name

    base_sheet = wb["基础信息"] if "基础信息" in wb.sheetnames else wb[wb.sheetnames[0]]
    base = read_key_value_sheet(base_sheet)

    module_rows = read_table_sheet(wb["模块规划"]) if "模块规划" in wb.sheetnames else []
    activity_rows = read_table_sheet(wb["主要活动"]) if "主要活动" in wb.sheetnames else []
    venue_rows = read_table_sheet(wb["场地与资源"]) if "场地与资源" in wb.sheetnames else []
    department_rows = read_table_sheet(wb["责任单位与联系人"]) if "责任单位与联系人" in wb.sheetnames else []
    supplier_rows = read_table_sheet(wb["供应商"]) if "供应商" in wb.sheetnames else []
    task_rows = read_table_sheet(wb["前期任务"]) if "前期任务" in wb.sheetnames else []
    risk_rows = read_table_sheet(wb["风险事项"]) if "风险事项" in wb.sheetnames else []
    agent_rows = read_table_sheet(wb["agent接入口"]) if "agent接入口" in wb.sheetnames else []

    recommended_modules = unique_list(row.get("模块名称") for row in module_rows)
    module_details = []
    for row in module_rows:
        name = normalize_text(row.get("模块名称"))
        if not name:
            continue
        module_details.append(
            {
                "name": name,
                "description": normalize_text(row.get("模块说明")),
                "startDate": normalize_date(row.get("开始日期")),
                "endDate": normalize_date(row.get("结束日期")),
                "owner": normalize_text(row.get("负责人")),
                "remark": normalize_text(row.get("备注")),
            }
        )

    activity_details = []
    for row in activity_rows:
        title = normalize_text(row.get("活动名称"))
        if not title:
            continue
        activity_details.append(
            {
                "name": title,
                "date": normalize_date(row.get("日期")),
                "time": normalize_text(row.get("时间")),
                "location": normalize_text(row.get("场地")),
                "module": normalize_text(row.get("对应模块")),
                "owner": normalize_text(row.get("负责人")),
                "assistant": normalize_text(row.get("协助人")),
                "priority": normalize_text(row.get("优先级")),
                "status": normalize_text(row.get("状态")),
                "remark": normalize_text(row.get("备注")),
            }
        )

    venue_details = []
    for row in venue_rows:
        venue = normalize_text(row.get("场地/资源"))
        if not venue:
            continue
        venue_details.append(
            {
                "name": venue,
                "purpose": normalize_text(row.get("用途")),
                "activity": normalize_text(row.get("所属活动")),
                "department": normalize_text(row.get("责任单位")),
                "owner": normalize_text(row.get("负责人")),
                "contact": normalize_text(row.get("联系方式")),
                "status": normalize_text(row.get("状态")),
                "remark": normalize_text(row.get("备注")),
            }
        )

    department_details = []
    for row in department_rows:
        unit = normalize_text(row.get("单位"))
        contact = normalize_text(row.get("联系人"))
        if not unit and not contact:
            continue
        department_details.append(
            {
                "unit": unit,
                "role": normalize_text(row.get("职能")),
                "contact": contact,
                "mobile": normalize_text(row.get("手机")),
                "feishuId": normalize_text(row.get("飞书 ID")),
                "matter": normalize_text(row.get("对接事项")),
                "level": normalize_text(row.get("级别")),
                "remark": normalize_text(row.get("备注")),
            }
        )

    supplier_details = []
    for row in supplier_rows:
        supplier = normalize_text(row.get("供应商"))
        if not supplier:
            continue
        supplier_details.append(
            {
                "name": supplier,
                "service": normalize_text(row.get("服务内容")),
                "contact": normalize_text(row.get("联系人")),
                "phone": normalize_text(row.get("电话")),
                "contractStatus": normalize_text(row.get("合同状态")),
                "owner": normalize_text(row.get("对接负责人")),
                "remark": normalize_text(row.get("备注")),
            }
        )

    task_drafts = []
    for row in task_rows:
        task = normalize_text(row.get("任务内容"))
        if not task:
            continue
        task_drafts.append(
            {
                "content": task,
                "module": normalize_text(row.get("所属模块")),
                "owner": normalize_text(row.get("负责人")),
                "assistant": normalize_text(row.get("协助人")),
                "startTime": normalize_text(row.get("开始时间")),
                "dueTime": normalize_text(row.get("截止时间")),
                "priority": normalize_text(row.get("优先级")),
                "status": normalize_text(row.get("状态")),
                "remark": normalize_text(row.get("备注")),
            }
        )

    risk_items = []
    for row in risk_rows:
        risk = normalize_text(row.get("风险描述"))
        if not risk:
            continue
        risk_items.append(
            {
                "description": risk,
                "module": normalize_text(row.get("影响模块")),
                "probability": normalize_text(row.get("发生概率")),
                "impact": normalize_text(row.get("影响程度")),
                "plan": normalize_text(row.get("预案")),
                "owner": normalize_text(row.get("负责人")),
                "dueDate": normalize_text(row.get("截止时间")),
                "status": normalize_text(row.get("状态")),
            }
        )

    agent_integration = {}
    for row in agent_rows:
        field = normalize_text(row.get("字段"))
        value = normalize_text(row.get("填写内容"))
        if field and value:
            agent_integration[field] = value

    key_contacts = []
    for item in department_details:
        if item["contact"]:
            key_contacts.append(item["contact"])
    for item in supplier_details:
        if item["contact"]:
            key_contacts.append(item["contact"])
    key_contacts = unique_list(key_contacts)

    primary_venues = unique_list(
        [
            normalize_text(base.get("项目地点")),
            *(item["name"] for item in venue_details),
            *(item["location"] for item in activity_details),
        ]
    )

    primary_departments = unique_list(item["unit"] for item in department_details)

    meta = {
        "projectName": normalize_text(base.get("项目名称")),
        "shortName": normalize_text(base.get("项目简称")),
        "folderName": normalize_text(base.get("项目目录")) or folder_name,
        "projectCode": normalize_text(base.get("项目编码")),
        "projectType": normalize_text(base.get("项目类型")),
        "projectStatus": normalize_text(base.get("当前阶段")) or "planning",
        "city": normalize_text(base.get("主要城市")),
        "projectDescription": normalize_text(base.get("项目描述")),
        "projectManager": normalize_text(base.get("项目经理")),
        "businessContact": normalize_text(base.get("业务对接人")),
        "dateRange": {
            "start": normalize_date(base.get("项目开始日期")),
            "end": normalize_date(base.get("项目结束日期")),
            "peakWindowStart": normalize_date(base.get("核心活动高峰开始")),
            "peakWindowEnd": normalize_date(base.get("核心活动高峰结束")),
        },
        "sourceDocument": {
            "fileName": normalize_text(base.get("来源文档文件名")) or file_name,
            "sheetName": normalize_text(base.get("来源文档表名")) or "基础信息",
            "title": normalize_text(base.get("来源文档标题")) or "项目前期录入模板",
        },
        "summary": {
            "activityCount": len(activity_details),
            "taskItemCount": len(task_drafts),
            "locationCount": len(primary_venues),
            "departmentCount": len(primary_departments),
            "contactCount": len(key_contacts),
        },
        "primaryActivities": unique_list(item["name"] for item in activity_details),
        "primaryVenues": primary_venues,
        "primaryDepartments": primary_departments,
        "keyContacts": key_contacts,
        "recommendedModules": recommended_modules,
        "moduleDetails": [
            {
                "name": item["name"],
                "description": item["description"],
                "startDate": item["startDate"],
                "endDate": item["endDate"],
                "owner": item["owner"],
                "remark": item["remark"],
            }
            for item in module_details
        ],
        "activityDetails": activity_details,
        "venueDetails": venue_details,
        "departmentContacts": department_details,
        "supplierDetails": supplier_details,
        "taskDrafts": task_drafts,
        "riskItems": risk_items,
        "agentIntegration": agent_integration,
        "intakeWorkbook": {
            "fileName": file_name,
            "sheetName": "填写说明",
        },
    }

    return meta


def main():
    parser = argparse.ArgumentParser(description="Convert a project intake xlsx into project.meta.json")
    parser.add_argument("workbook", help="Path to the intake workbook")
    parser.add_argument(
        "--output-meta",
        help="Write project.meta.json to this path. Defaults to the workbook folder.",
    )
    parser.add_argument(
        "--stdout",
        action="store_true",
        help="Print the JSON to stdout instead of writing a file.",
    )
    args = parser.parse_args()

    workbook_path = Path(args.workbook).expanduser().resolve()
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    meta = build_meta(workbook_path)
    meta_json = json.dumps(meta, ensure_ascii=False, indent=2)

    if args.stdout:
        print(meta_json)
        return

    output_path = Path(args.output_meta).expanduser().resolve() if args.output_meta else workbook_path.parent / "project.meta.json"
    output_path.write_text(meta_json + "\n", encoding="utf-8")
    print(f"Written: {output_path}")


if __name__ == "__main__":
    main()
