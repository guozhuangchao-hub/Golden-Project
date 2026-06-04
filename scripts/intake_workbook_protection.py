#!/usr/bin/env python3
from copy import copy

from openpyxl import load_workbook


YELLOW_FILL_RGBS = {"00FFF2CC", "FFF2CC"}
PASSWORD = "yhgg"


def _fill_rgb(cell):
    fill = cell.fill
    if fill is None or fill.patternType != "solid":
        return ""

    color = fill.fgColor
    if color is None:
        return ""

    if color.type == "rgb" and color.rgb:
        return str(color.rgb).upper()
    if color.type == "indexed" and color.indexed is not None:
        return f"indexed:{color.indexed}"
    return ""


def _protect_sheet(ws, *, unlock_input_cells=None, lock_cells=None, protect_all=False):
    unlock_input_cells = unlock_input_cells or set()
    lock_cells = lock_cells or set()

    for row in ws.iter_rows():
        for cell in row:
            protection = copy(cell.protection)
            protection.locked = True

            if protect_all:
                cell.protection = protection
                continue

            if cell.coordinate in lock_cells:
                cell.protection = protection
                continue

            if cell.coordinate in unlock_input_cells:
                protection.locked = False
                cell.protection = protection
                continue

            if _fill_rgb(cell) in YELLOW_FILL_RGBS:
                protection.locked = False

            cell.protection = protection

    ws.protection.sheet = True
    ws.protection.set_password(PASSWORD)


def protect_intake_workbook(workbook_path, *, mode="template"):
    wb = load_workbook(workbook_path)

    for ws in wb.worksheets:
        if ws.title == "填写说明":
            _protect_sheet(ws, protect_all=True)
            continue

        if ws.title == "原始事项清单":
            _protect_sheet(ws, protect_all=True)
            continue

        lock_cells = set()
        if mode == "project" and ws.title == "基础信息":
            lock_cells.update(
                {
                    "B5",
                    "B6",
                    "B7",
                    "B19",
                    "B20",
                    "B21",
                }
            )

        _protect_sheet(ws, lock_cells=lock_cells)

    wb.save(workbook_path)
    return workbook_path
