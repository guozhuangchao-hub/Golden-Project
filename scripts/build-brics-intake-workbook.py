#!/usr/bin/env python3
from copy import copy
import json
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from intake_workbook_protection import protect_intake_workbook


ROOT_DIR = Path(__file__).resolve().parents[1]
TEMPLATE_PATH = ROOT_DIR / "项目列表" / "项目模板" / "前期录入模板.xlsx"
SOURCE_XLSX = ROOT_DIR / "项目列表" / "26年5月金砖项目" / "2026金砖国家新工业革命伙伴关系论坛.xlsx"
META_PATH = ROOT_DIR / "项目列表" / "26年5月金砖项目" / "project.meta.json"
OUTPUT_PATH = ROOT_DIR / "项目列表" / "26年5月金砖项目" / "2026金砖国家新工业革命伙伴关系论坛-前期录入.xlsx"


def norm(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return str(value).strip()


def excel_date(value, epoch):
    if value in (None, ""):
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, (int, float)):
        try:
            return from_excel(value, epoch=epoch).date().isoformat()
        except Exception:
            return str(value)
    return str(value)


def load_meta():
    return json.loads(META_PATH.read_text(encoding="utf-8"))


def clone_cell_style(src, dst):
    dst.font = copy(src.font)
    dst.fill = copy(src.fill)
    dst.border = copy(src.border)
    dst.alignment = copy(src.alignment)
    dst.number_format = src.number_format
    dst.protection = copy(src.protection)


def write_row(ws, row_idx, values, style_row_idx=5):
    max_col = max(len(values), ws.max_column)
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        if col_idx <= len(values):
            cell.value = values[col_idx - 1]
        src = ws.cell(row=style_row_idx, column=col_idx)
        clone_cell_style(src, cell)


def clear_rows(ws, start_row, end_row):
    for row_idx in range(start_row, end_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).value = None


def infer_module(activity, need, status):
    text = f"{norm(activity)} {norm(need)} {norm(status)}"
    rules = [
        (["新闻发布", "宣传", "发布会", "媒体"], "新闻发布与宣传材料"),
        (["开幕式", "主论坛", "创新大赛", "理事会", "成果发布"], "主论坛与开幕式"),
        (["分论坛", "交流会", "产业对接", "展览", "巡馆", "决赛"], "展览与产业对接"),
        (["会见", "休息室", "外宾", "接待", "嘉宾"], "嘉宾接待与会见"),
        (["晚宴", "早餐叙", "餐叙", "欢迎晚宴", "午餐"], "餐叙与晚宴"),
        (["翻译", "同传", "俄语", "英语", "语种"], "翻译与同传安排"),
        (["文稿", "主持词", "致辞", "祝酒辞", "发言稿", "问答"], "文稿与发言材料"),
        (["会场", "座位", "布置", "音响", "音乐", "视频", "摄影", "安检", "入场", "引导", "车辆", "转场", "物料"], "会场布置与会务执行"),
        (["通知", "报名", "名单", "引导", "人员"], "通知与人员引导"),
    ]
    for keywords, module in rules:
        if any(keyword in text for keyword in keywords):
            return module
    return "会场布置与会务执行"


def derive_priority(status):
    text = norm(status)
    if any(key in text for key in ["动态调整", "待确认", "准备中", "初稿"]):
        return "高"
    if any(key in text for key in ["制定中", "收集中"]):
        return "中"
    if "已完成" in text:
        return "低"
    return "中"


def summarize_activity(items):
    statuses = [norm(item["status"]) for item in items if norm(item["status"])]
    if not statuses:
        return "待确认"
    if all("已完成" in s for s in statuses):
        return "已完成"
    if any("动态调整" in s or "待确认" in s or "准备中" in s for s in statuses):
        return "执行中"
    if any("初稿" in s or "制定中" in s or "收集中" in s for s in statuses):
        return "执行中"
    return "执行中"


def parse_source():
    wb = load_workbook(SOURCE_XLSX, data_only=True)
    ws = wb.active

    groups = []
    current = None
    for r in range(4, ws.max_row + 1):
        seq = ws.cell(r, 1).value
        if seq is not None:
            raw_time = ws.cell(r, 2).value
            time_text = norm(raw_time)
            date_text = ""
            clock_text = ""
            if isinstance(raw_time, (int, float)):
                date_text = excel_date(raw_time, wb.epoch)
            elif isinstance(raw_time, str):
                parts = [part.strip() for part in raw_time.split("\n") if part.strip()]
                if parts:
                    date_text = parts[0]
                if len(parts) > 1:
                    clock_text = " ".join(parts[1:])
                if not date_text:
                    date_text = time_text
            current = {
                "seq": int(seq),
                "date": date_text,
                "time": clock_text,
                "time_raw": time_text,
                "location": norm(ws.cell(r, 3).value),
                "activity": norm(ws.cell(r, 4).value),
                "items": [],
            }
            groups.append(current)

        if current is None:
            continue

        need = norm(ws.cell(r, 5).value)
        status = norm(ws.cell(r, 6).value)
        dept = norm(ws.cell(r, 7).value)
        contact = norm(ws.cell(r, 8).value)
        if need or status or dept or contact:
            current["items"].append(
                {
                    "need": need,
                    "status": status,
                    "dept": dept,
                    "contact": contact,
                }
            )

    return groups


def unique_preserve(values):
    seen = []
    for value in values:
        if value and value not in seen:
            seen.append(value)
    return seen


def fill_basic_info(ws, meta):
    mapping = {
        "项目名称": meta.get("projectName", ""),
        "项目简称": meta.get("shortName", ""),
        "项目编码": meta.get("projectCode", ""),
        "项目类型": meta.get("projectType", ""),
        "当前阶段": meta.get("projectStatus", ""),
        "主要城市": meta.get("city", ""),
        "项目地点": meta.get("primaryVenues", [""])[0] if meta.get("primaryVenues") else meta.get("city", ""),
        "项目经理": meta.get("projectManager", "待补充"),
        "业务对接人": meta.get("businessContact", "待补充"),
        "项目描述": meta.get("projectDescription", ""),
        "项目开始日期": meta.get("dateRange", {}).get("start", ""),
        "项目结束日期": meta.get("dateRange", {}).get("end", ""),
        "核心活动高峰开始": meta.get("dateRange", {}).get("peakWindowStart", ""),
        "核心活动高峰结束": meta.get("dateRange", {}).get("peakWindowEnd", ""),
        "来源文档文件名": meta.get("sourceDocument", {}).get("fileName", ""),
        "来源文档表名": meta.get("sourceDocument", {}).get("sheetName", ""),
        "来源文档标题": meta.get("sourceDocument", {}).get("title", ""),
        "备注": "本表根据原始事项清单自动整理，供项目经理前期录入与 agent 读取。",
    }
    for row in range(5, ws.max_row + 1):
        field = norm(ws.cell(row, 1).value)
        if field in mapping:
            ws.cell(row, 2).value = mapping[field]


def fill_module_sheet(ws, meta):
    modules = meta.get("recommendedModules", [])
    details = {item.get("name"): item for item in meta.get("moduleDetails", []) if item.get("name")}
    clear_rows(ws, 5, ws.max_row)
    for idx, module_name in enumerate(modules, start=5):
        detail = details.get(module_name, {})
        write_row(
            ws,
            idx,
            [
                idx - 4,
                module_name,
                norm(detail.get("description", "")),
                norm(detail.get("startDate", "")),
                norm(detail.get("endDate", "")),
                norm(detail.get("owner", "")),
                "来源：金砖项目原始事项与模板映射",
            ],
        )


def fill_activity_sheet(ws, groups):
    clear_rows(ws, 5, ws.max_row)
    for row_idx, group in enumerate(groups, start=5):
        items = group["items"]
        contacts = unique_preserve([item["contact"] for item in items if item["contact"]])
        row = [
            group["seq"],
            group["activity"],
            group["date"],
            group["time"],
            group["location"],
            infer_module(group["activity"], " ".join(item["need"] for item in items), " ".join(item["status"] for item in items)),
            items[0]["dept"] if items else "",
            contacts[0] if contacts else "",
            derive_priority(" ".join(item["status"] for item in items)),
            summarize_activity(items),
            f"准备事项：{'；'.join(item['need'] for item in items if item['need'])}",
        ]
        write_row(ws, row_idx, row)


def fill_resource_sheet(ws, groups):
    resources = []
    for group in groups:
        if group["location"]:
            contacts = unique_preserve([item["contact"] for item in group["items"] if item["contact"]])
            resources.append(
                (
                    group["location"],
                    "活动场地/资源",
                    group["activity"],
                    group["items"][0]["dept"] if group["items"] else "",
                    contacts[0] if contacts else "",
                    "",
                    "待确认",
                    "来源：原始事项清单"
                )
            )
    clear_rows(ws, 5, ws.max_row)
    for idx, item in enumerate(unique_preserve(resources), start=5):
        write_row(ws, idx, [idx - 4, *item])


def fill_contacts_sheet(ws, groups):
    contacts = []
    for group in groups:
        for item in group["items"]:
            if item["dept"] or item["contact"]:
                contacts.append(
                    (
                        item["dept"],
                        "事项对接",
                        item["contact"],
                        "",
                        "",
                        group["activity"],
                        "核心" if "局" in item["dept"] or "中心" in item["dept"] else "普通",
                        item["status"],
                    )
                )
    clear_rows(ws, 5, ws.max_row)
    for idx, item in enumerate(unique_preserve(contacts), start=5):
        write_row(ws, idx, [idx - 4, *item])


def fill_supplier_sheet(ws):
    clear_rows(ws, 5, ws.max_row)
    sample = [
        (1, "待补充", "物料/搭建/翻译/车辆等合作方后续补录", "", "", "未签约", "待指定", "当前原始清单未单列供应商"),
    ]
    for row_idx, row in enumerate(sample, start=5):
        write_row(ws, row_idx, row)


def fill_task_sheet(ws, groups):
    tasks = []
    for group in groups:
        module = infer_module(group["activity"], "", "")
        for item in group["items"]:
            tasks.append(
                (
                    item["need"],
                    module,
                    item["dept"],
                    item["contact"],
                    "",
                    "",
                    "高" if derive_priority(item["status"]) == "高" else "中",
                    item["status"] or "待确认",
                    group["activity"],
                )
            )
    clear_rows(ws, 5, ws.max_row)
    for idx, row in enumerate(tasks, start=5):
        write_row(ws, idx, [idx - 4, *row])


def fill_risk_sheet(ws, groups):
    risks = []
    for group in groups:
        for item in group["items"]:
            if any(key in item["status"] for key in ["待确认", "动态调整", "准备中", "初稿", "制定中", "收集中"]):
                risks.append(
                    (
                        f"{group['activity']} - {item['need']}",
                        infer_module(group["activity"], item["need"], item["status"]),
                        "高" if any(key in item["status"] for key in ["待确认", "动态调整", "准备中"]) else "中",
                        "高" if any(key in item["status"] for key in ["待确认", "动态调整", "准备中"]) else "中",
                        f"优先确认：{item['need']} 的责任单位、时间与名单。",
                        item["dept"],
                        group["date"] or group["time_raw"],
                        "待确认",
                    )
                )
    # de-duplicate
    risks = unique_preserve(risks)[:12]
    clear_rows(ws, 5, ws.max_row)
    for idx, row in enumerate(risks, start=5):
        write_row(ws, idx, [idx - 4, *row])


def fill_agent_sheet(ws):
    mapping = {
        "可接入 agent": "Hermes / 其他",
        "接入方式": "webhook / API",
        "读取范围": "项目基础信息、模块、活动、任务、风险、消息",
        "写回范围": "任务创建、状态更新、日报写入、通知推送",
        "审核方式": "项目经理确认",
        "备注": "预留给后续其他 agent 接入维护",
    }
    for row in range(5, ws.max_row + 1):
        field = norm(ws.cell(row, 1).value)
        if field in mapping:
            ws.cell(row, 2).value = mapping[field]


def build_output():
    meta = load_meta()
    groups = parse_source()

    wb = load_workbook(TEMPLATE_PATH)
    fill_basic_info(wb["基础信息"], meta)
    fill_module_sheet(wb["模块规划"], meta)
    fill_activity_sheet(wb["主要活动"], groups)
    fill_resource_sheet(wb["场地与资源"], groups)
    fill_contacts_sheet(wb["责任单位与联系人"], groups)
    fill_supplier_sheet(wb["供应商"])
    fill_task_sheet(wb["前期任务"], groups)
    fill_risk_sheet(wb["风险事项"], groups)
    fill_agent_sheet(wb["agent接入口"])

    # Add a normalized raw-data sheet for traceability.
    if "原始事项清单" in wb.sheetnames:
        del wb["原始事项清单"]
    raw = wb.create_sheet("原始事项清单")
    headers = ["序号", "时间", "地点", "主要活动", "需准备事项", "进展情况", "责任单位/部门", "联系人", "对应模块", "备注"]
    for col_idx, header in enumerate(headers, start=1):
        raw.cell(row=1, column=col_idx, value=header)
    source_rows = []
    for group in groups:
        module = infer_module(group["activity"], "", "")
        for item in group["items"]:
            source_rows.append(
                [
                    group["seq"],
                    group["date"] or group["time_raw"],
                    group["location"],
                    group["activity"],
                    item["need"],
                    item["status"],
                    item["dept"],
                    item["contact"],
                    module,
                    "自动从原始事项清单拆分",
                ]
            )
    for idx, row in enumerate(source_rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            raw.cell(row=idx, column=col_idx, value=value)
            clone_cell_style(wb["主要活动"].cell(row=5, column=min(col_idx, wb["主要活动"].max_column)), raw.cell(row=idx, column=col_idx))
    raw.freeze_panes = "A2"

    wb.save(OUTPUT_PATH)
    protect_intake_workbook(OUTPUT_PATH, mode="project")
    print(f"Written: {OUTPUT_PATH}")


if __name__ == "__main__":
    build_output()
