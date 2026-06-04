#!/usr/bin/env python3
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation

from intake_workbook_protection import protect_intake_workbook


ROOT_DIR = Path(__file__).resolve().parents[1]
TEMPLATE_PATH = ROOT_DIR / "项目列表" / "项目模板" / "前期录入模板.xlsx"

FONT_NAME = "Arial"
TITLE_FILL = "1F4E78"
SECTION_FILL = "D9E2F3"
HEADER_FILL = "BDD7EE"
INPUT_FILL = "FFF2CC"
NOTE_FILL = "F3F6FA"
WHITE = "FFFFFF"
GRID = "D9D9D9"

thin_gray = Side(style="thin", color=GRID)
border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)


def style_cell(cell, *, bold=False, color="000000", fill=None, size=11, center=False):
    cell.font = Font(name=FONT_NAME, size=size, bold=bold, color=color)
    cell.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center",
        wrap_text=True,
    )
    cell.border = border
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)


def merge_title(ws, title, end_col="F"):
    ws.merge_cells(f"A1:{end_col}1")
    ws["A1"] = title
    style_cell(ws["A1"], bold=True, color=WHITE, fill=TITLE_FILL, size=14, center=True)
    ws.row_dimensions[1].height = 24


def add_note(ws, text, end_col="F"):
    ws.merge_cells(f"A2:{end_col}2")
    ws["A2"] = text
    style_cell(ws["A2"], fill=NOTE_FILL)
    ws.row_dimensions[2].height = 34


def apply_col_widths(ws, widths):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def create_key_value_sheet(wb):
    ws = wb.create_sheet("基础信息")
    merge_title(ws, "项目基础信息录入")
    add_note(ws, "先填写这些字段，agent 会据此生成项目基础结构。黄色单元格为填写区。", "C")

    headers = ["字段", "填写内容", "说明"]
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=idx, value=header)
        style_cell(cell, bold=True, fill=HEADER_FILL, center=True)

    rows = [
        ("项目名称", "", "活动全称"),
        ("项目简称", "", "便于后台和群里快速识别"),
        ("项目编码", "", "建议唯一，例如 2026JZGJXGYGMHBGXLT20260602YHGG，规则为“项目名称大写缩写 + 创建日期 + YHGG”"),
        ("项目类型", "", "例如 forum_event / summit / launch"),
        ("当前阶段", "planning", "可选 planning / active / completed / cancelled"),
        ("主要城市", "", "项目举办城市"),
        ("项目地点", "", "活动主要落点或场地名称"),
        ("项目经理", "", "负责项目整体推进的人"),
        ("业务对接人", "", "客户或主办方对接人"),
        ("项目描述", "", "一句话概括项目目标与范围"),
        ("项目开始日期", "", "格式建议 YYYY-MM-DD"),
        ("项目结束日期", "", "格式建议 YYYY-MM-DD"),
        ("核心活动高峰开始", "", "最忙时间段开始"),
        ("核心活动高峰结束", "", "最忙时间段结束"),
        ("来源文档文件名", "", "原始 Excel / PDF / Word 文件名"),
        ("来源文档表名", "", "若原始文件是 Excel，请填写 sheet 名"),
        ("来源文档标题", "", "原始文件标题或清单名称"),
        ("备注", "", "可写补充说明"),
    ]

    for row_idx, (field, value, note) in enumerate(rows, start=5):
        ws.cell(row=row_idx, column=1, value=field)
        style_cell(ws.cell(row=row_idx, column=1), fill=SECTION_FILL, bold=True)
        value_cell = ws.cell(row=row_idx, column=2, value=value)
        style_cell(value_cell, fill=INPUT_FILL)
        ws.cell(row=row_idx, column=3, value=note)
        style_cell(ws.cell(row=row_idx, column=3), fill=NOTE_FILL)

    apply_col_widths(ws, {"A": 22, "B": 38, "C": 46})
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False
    return ws


def create_table_sheet(wb, title, note, headers, sample_rows=12, col_widths=None, validation_map=None):
    ws = wb.create_sheet(title)
    merge_title(ws, title)
    add_note(ws, note)

    header_row = 4
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=idx, value=header)
        style_cell(cell, bold=True, fill=HEADER_FILL, center=True)

    for row in range(header_row + 1, header_row + sample_rows + 1):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col, value="")
            style_cell(cell, fill=INPUT_FILL if col != 1 else SECTION_FILL)

        ws.cell(row=row, column=1, value=row - header_row)
        style_cell(ws.cell(row=row, column=1), fill=SECTION_FILL, bold=True, center=True)

    if col_widths:
        apply_col_widths(ws, col_widths)

    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = f"A{header_row}:{chr(64 + len(headers))}{header_row + sample_rows}"
    ws.sheet_view.showGridLines = False

    if validation_map:
        for col_letter, formula1 in validation_map.items():
            dv = DataValidation(type="list", formula1=formula1, allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f"{col_letter}{header_row + 1}:{col_letter}{header_row + sample_rows}")

    return ws


def build_workbook():
    wb = Workbook()
    default_ws = wb.active
    wb.remove(default_ws)

    overview = wb.create_sheet("填写说明")
    merge_title(overview, "Golden Project 项目前期录入模板", "G")
    add_note(overview, "建议项目经理先填「基础信息」和「模块规划」，然后再补「主要活动」「前期任务」「风险事项」。", "G")
    overview["A4"] = "使用规则"
    style_cell(overview["A4"], bold=True, fill=HEADER_FILL)
    overview["B4"] = "说明"
    style_cell(overview["B4"], bold=True, fill=HEADER_FILL)

    rules = [
        ("1", "每个项目一个工作簿，放到对应项目文件夹内。"),
        ("2", "黄色单元格为填写区，灰蓝色单元格为字段或序号。"),
        ("3", "日期建议使用 YYYY-MM-DD。"),
        ("4", "模块、活动、任务、风险尽量一行一条，便于 agent 读取。"),
        ("5", "填写完成后，运行 `npm run sync:projects`，agent 会先生成或刷新 `project.meta.json`，再同步到数据库。"),
        ("6", "如果项目后续要接飞书、Hermes 或其他 agent，请在「agent接入口」里补齐信息。"),
    ]
    for idx, (no, text) in enumerate(rules, start=5):
        overview.cell(row=idx, column=1, value=no)
        style_cell(overview.cell(row=idx, column=1), fill=SECTION_FILL, bold=True, center=True)
        overview.cell(row=idx, column=2, value=text)
        style_cell(overview.cell(row=idx, column=2), fill=INPUT_FILL)

    overview["D4"] = "Sheet"
    style_cell(overview["D4"], bold=True, fill=HEADER_FILL, center=True)
    overview["E4"] = "用途"
    style_cell(overview["E4"], bold=True, fill=HEADER_FILL, center=True)
    mapping = [
        ("基础信息", "项目名称、编码、时间范围、项目描述等"),
        ("模块规划", "推荐模块与模块时间"),
        ("主要活动", "活动节点、负责人、场地、优先级"),
        ("场地与资源", "场地、设备、资源与责任单位"),
        ("责任单位与联系人", "单位和联络人清单"),
        ("供应商", "供应商与合作方信息"),
        ("前期任务", "启动前任务清单"),
        ("风险事项", "风险、预案和跟踪"),
        ("agent接入口", "其他 agent 的接入信息"),
    ]
    for idx, (sheet_name, purpose) in enumerate(mapping, start=5):
        overview.cell(row=idx, column=4, value=sheet_name)
        style_cell(overview.cell(row=idx, column=4), fill=SECTION_FILL, bold=True)
        overview.cell(row=idx, column=5, value=purpose)
        style_cell(overview.cell(row=idx, column=5), fill=INPUT_FILL)

    apply_col_widths(overview, {"A": 8, "B": 52, "D": 18, "E": 36, "F": 14, "G": 14})
    overview.freeze_panes = "A5"
    overview.sheet_view.showGridLines = False

    create_key_value_sheet(wb)

    create_table_sheet(
        wb,
        "模块规划",
        "建议把活动拆成 8-12 个模块。模块会同步成 project_modules，后续任务归属会优先挂在这里。",
        ["序号", "模块名称", "模块说明", "开始日期", "结束日期", "负责人", "备注"],
        sample_rows=12,
        col_widths={"A": 8, "B": 24, "C": 34, "D": 14, "E": 14, "F": 14, "G": 22},
    )

    create_table_sheet(
        wb,
        "主要活动",
        "按时间先后填写。每一行建议只描述一个活动节点或一个关键事件。",
        ["序号", "活动名称", "日期", "时间", "场地", "对应模块", "负责人", "协助人", "优先级", "状态", "备注"],
        sample_rows=16,
        col_widths={"A": 8, "B": 24, "C": 14, "D": 16, "E": 20, "F": 18, "G": 14, "H": 14, "I": 10, "J": 12, "K": 22},
        validation_map={
            "I": '"高,中,低,紧急"',
            "J": '"待确认,已确认,执行中,已完成,已逾期,已取消"',
        },
    )

    create_table_sheet(
        wb,
        "场地与资源",
        "场地、设备、车辆、物资、系统账号等都可以放进来，便于项目经理统一查看。",
        ["序号", "场地/资源", "用途", "所属活动", "责任单位", "负责人", "联系方式", "状态", "备注"],
        sample_rows=12,
        col_widths={"A": 8, "B": 22, "C": 18, "D": 22, "E": 16, "F": 14, "G": 16, "H": 12, "I": 22},
        validation_map={
            "H": '"待确认,已确认,使用中,已完成,已取消"',
        },
    )

    create_table_sheet(
        wb,
        "责任单位与联系人",
        "用于沉淀项目各分部、执行方和关键联系人，后续 agent 会优先从这里找责任链。",
        ["序号", "单位", "职能", "联系人", "手机", "飞书 ID", "对接事项", "级别", "备注"],
        sample_rows=16,
        col_widths={"A": 8, "B": 18, "C": 16, "D": 12, "E": 14, "F": 18, "G": 18, "H": 10, "I": 20},
        validation_map={
            "H": '"核心,普通,外部,临时"',
        },
    )

    create_table_sheet(
        wb,
        "供应商",
        "如果后续有搭建、物料、车辆、翻译、安保等供应商，直接在这里登记。",
        ["序号", "供应商", "服务内容", "联系人", "电话", "合同状态", "对接负责人", "备注"],
        sample_rows=12,
        col_widths={"A": 8, "B": 20, "C": 18, "D": 12, "E": 14, "F": 12, "G": 14, "H": 20},
        validation_map={
            "F": '"未签约,洽谈中,已签约,已完成,已取消"',
        },
    )

    create_table_sheet(
        wb,
        "前期任务",
        "项目启动前的准备事项。后续 agent 也可以把这里的内容转成系统任务。",
        ["序号", "任务内容", "所属模块", "负责人", "协助人", "开始时间", "截止时间", "优先级", "状态", "备注"],
        sample_rows=20,
        col_widths={"A": 8, "B": 28, "C": 18, "D": 14, "E": 14, "F": 16, "G": 16, "H": 10, "I": 12, "J": 22},
        validation_map={
            "H": '"高,中,低,紧急"',
            "I": '"待确认,已确认,执行中,已完成,已逾期,已取消"',
        },
    )

    create_table_sheet(
        wb,
        "风险事项",
        "把可能影响时间、资源和交付的事项提前记录，便于项目经理盯风险。",
        ["序号", "风险描述", "影响模块", "发生概率", "影响程度", "预案", "负责人", "截止时间", "状态"],
        sample_rows=12,
        col_widths={"A": 8, "B": 24, "C": 18, "D": 10, "E": 10, "F": 22, "G": 14, "H": 16, "I": 12},
        validation_map={
            "D": '"低,中,高"',
            "E": '"低,中,高"',
            "I": '"待确认,已确认,处理中,已缓解,已关闭"',
        },
    )

    agent = wb.create_sheet("agent接入口")
    merge_title(agent, "agent 接入口说明")
    add_note(agent, "这里给 Hermes 或其他 agent 预留接入信息。项目经理先填读取范围和写回范围，后续平台会据此开放权限。", "G")

    headers = ["字段", "填写内容", "说明"]
    for idx, header in enumerate(headers, start=1):
        cell = agent.cell(row=4, column=idx, value=header)
        style_cell(cell, bold=True, fill=HEADER_FILL, center=True)

    agent_rows = [
        ("可接入 agent", "Hermes / 其他", "后续可由其他 agent 接入维护"),
        ("接入方式", "webhook / API", "系统预留通用入口"),
        ("读取范围", "", "例如：项目、任务、消息、提案、日报"),
        ("写回范围", "", "例如：任务创建、状态更新、日报写入"),
        ("审核方式", "项目经理确认", "项目经理确认 / 自动确认 / 混合"),
        ("备注", "", "可补充 agent 的角色边界"),
    ]
    for row_idx, (field, value, note) in enumerate(agent_rows, start=5):
        agent.cell(row=row_idx, column=1, value=field)
        style_cell(agent.cell(row=row_idx, column=1), fill=SECTION_FILL, bold=True)
        agent.cell(row=row_idx, column=2, value=value)
        style_cell(agent.cell(row=row_idx, column=2), fill=INPUT_FILL)
        agent.cell(row=row_idx, column=3, value=note)
        style_cell(agent.cell(row=row_idx, column=3), fill=NOTE_FILL)

    apply_col_widths(agent, {"A": 18, "B": 28, "C": 44})
    agent.sheet_view.showGridLines = False
    agent.freeze_panes = "A5"

    return wb


def main():
    TEMPLATE_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = build_workbook()
    wb.save(TEMPLATE_PATH)
    protect_intake_workbook(TEMPLATE_PATH, mode="template")
    print(f"Template written: {TEMPLATE_PATH}")


if __name__ == "__main__":
    main()
